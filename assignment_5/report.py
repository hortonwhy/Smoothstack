import os
import re
import pandas as pd
from loguru import logger
from openpyxl import load_workbook


class Report:
    def __init__(self, dirpath="./"):
        self.dirpath = dirpath
        self.filenames = []
        logger.add("out.log", backtrace=True, diagnose=True)
        self.summary_df = None
        self.voc_df = None

    def get_filenames(self):
        return self.filenames

    def load_filenames(self):
        logger.debug("Checking for valid files in directory: {}", self.dirpath)
        try:
            if isinstance(self.dirpath, (str)):
                listdir = os.listdir(self.dirpath)
            else:
                error = "Type of directory path must be a string"
                logger.error("{}: " + str(self.dirpath), error)
                return

        except OSError as oe:
            logger.error("OS error: {}", format(oe))
            return

        for filepath in listdir:
            if re.search(r"([a-zA-Z]:(\\w+)*\\[a-zA-Z0_9]+)?.xlsx", filepath):
                self.filenames.append(filepath)
        logger.debug("found valid files: {}", self.filenames)

    def load_sheets(self):
        for workbook in self.filenames:
            logger.debug("Opening {}", workbook)
            wb = load_workbook(workbook)

            ws = wb[wb.sheetnames[0]]  # Summary Rolling MoM
            self.summary_df = (pd.DataFrame(list(ws.values)), workbook)
            logger.debug("loaded Summary Rolling MoM for {}", workbook)
            self.log_summary(workbook)

            ws = wb[wb.sheetnames[1]]  # VOC Rolling MoM
            self.voc_df = (pd.DataFrame(list(ws.values)), workbook)
            logger.debug("loaded VOC Rolling MoM for {}", workbook)
            self.log_promoter_score(workbook)

    def format_date(self, date):
        # format Date Block to easily match with Workbook style
        month_to_date = {
            "jan": "01",
            "feb": "02",
            "mar": "03",
            "apr": "04",
            "may": "05",
            "jun": "06",
            "jul": "07",
            "aug": "08",
            "sep": "09",
            "oct": "10",
            "nov": "11",
            "dec": "12",
        }
        date = date.split("_")[3:]
        date_text = date[-2]  # eg. march
        date[0] = date[0][0:3]
        date[0] = date[0][0] + date[0][1:]
        date[0] = month_to_date[date[0]]
        date[1] = date[1].split(".")[0]
        return (date, date_text)

    def find_df_row(self, df, date, date_text):
        # match a date row with given date or date_text
        # eg. date is '01', '2018', eg. date_text is "march"
        for index, row in df.iterrows():
            if row[0] is not None:
                date_row = str(row[0]).split(" ")[0]  # 2017-03-01
                date_row = date_row.split("-")[0:2][::-1]  # ['01', '2018']
                if date_row == date or str(row[0]).lower() == date_text:
                    return row  # found row

    def log_summary(self, workbook):
        (date, date_text) = self.format_date(self.summary_df[1])
        logger.info("Logging Summary for {} from {}", date, workbook)
        df = self.summary_df[0]
        row = self.find_df_row(df, date, date_text)

        logger.info("Date: {}-{}", date[0], date[1])
        logger.info("Calls offered: {}", row[1])
        logger.info("Abandon after 30s: {}%", row[2] * 100)
        logger.info("FCR: {}%", row[3] * 100)
        logger.info("DSAT: {}%", row[4] * 100)
        logger.info("CSAT: {}%", row[5] * 100)

    def log_promoter_score(self, workbook):
        (date, date_text) = self.format_date(self.summary_df[1])
        logger.info("Logging Promoter Score for {} from {}", date, workbook)
        df = self.voc_df[0]
        df = df.T  # transpose
        row = self.find_df_row(df, date, date_text)
        logger.info("Date: {}-{}", date[0], date[1])
        logger.info("Base Size: {}", row[2])
        logger.info(
            "Promoters: {} {}%",
            str(row[3]) + " Good" if row[3] > 200 else str(row[3]) + " Bad",
            row[4] * 100,
        )
        logger.info(
            "Passives: {} {}%",
            str(row[5]) + " Good" if row[5] > 100 else str(row[5]) + " Bad",
            row[6] * 100,
        ),
        logger.info(
            "Detractors: {} {}%",
            str(row[7]) + " Good" if row[7] > 100 else str(row[7]) + " Bad",
            row[8] * 100,
        ),

        logger.info("Overall NPS: AARP Total {}%", row[12] * 100)
        logger.info("Sat with Agent: AARP Total {}%", row[15] * 100)
        logger.info("DSat with Agent: AARP Total {}%", row[18] * 100)
